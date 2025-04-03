import openpyxl
import matplotlib.pyplot as plt
from plotnine import *

print('Script is active')

x = r"C:\files\python scripts\polarization_curve_17_3.xlsx"
# x = input('Enter the path of the excel file: ')
title = input('Enter the title of the plot: ')
plot_matplotlib = input('Do you want to plot using matplotlib? (True/False): ')
save = input('Do you want to save the plot? (True/False): ')

def generate_polarization_curve(x, title, plot_nine = True, plot_matplotlib = False, save = False):
    print('hey')
    wb = openpyxl.load_workbook(x)
    ws = wb.active

    # Get the number of rows and columns
    print('Total number of rows: '+str(ws.max_row-1)+'. And total number of columns: '+str(ws.max_column))

    # Get the number of MFCs
    print('Total number of operating MFCs: '+str(ws['A1'].value))
    
    # Read the data from the excel file

    if ws['A1'].value>1:    
        resistance = [ws.cell(row=i,column=1).value for i in range(3,ws.max_row+1)]
        
        voltage = [[ws.cell(row=i,column=j).value for i in range(3,ws.max_row+1)] for j in range(2, 2+ws['A1'].value)]
        voltage[2], voltage[3] = voltage[3], voltage[2]

        current_density_raw = [[ws.cell(row=i,column=j).value for i in range(3,ws.max_row+1)] for j in range(2+ws['A1'].value, 2+2*ws['A1'].value)]
        current_density_raw[2], current_density_raw[3] = current_density_raw[3], current_density_raw[2]
        current_density_average = [sum(current_density_raw[i][j] for i in range(len(current_density_raw)))/ws['A1'].value for j in range(len(current_density_raw[0]))]


        power_density = [[ws.cell(row=i,column=j).value for i in range(3,ws.max_row+1)] for j in range(2+2*ws['A1'].value, 2+3*ws['A1'].value)]
        power_density[2], power_density[3] = power_density[3], power_density[2]
    
    else:
        resistance = [ws.cell(row=i,column=1).value for i in range(3,ws.max_row+1)]
        
        voltage = [ws.cell(row=i,column=2).value for i in range(3,ws.max_row+1)]

        current_density = [ws.cell(row=i,column=3).value for i in range(3,ws.max_row+1)]

        power_density = [ws.cell(row=i,column=4).value for i in range(3,ws.max_row+1)]

    # Generate the polarization curve
    if plot_matplotlib==True:
        fig, ax = plt.subplots()

        ax.set_xlabel('Current density (mA/m^2) -', fontsize = 'large', font = 'Times New Roman')
        ax.set_ylabel('Voltage (V)', fontsize = 'large', font = 'Times New Roman')
        for i in range(len(voltage)):
            ax.plot(current_density_average,voltage[i], label = 'MFC '+str(i+1))
        ax.tick_params(axis ='y', labelcolor = 'red') 
        
        ax2 = ax.twinx()
        ax2.set_ylabel('Power density (mW/m^2) --', fontsize = 'large', font = 'Times New Roman')
        for i in range(len(power_density)):
            ax2.plot(current_density_average,power_density[i], label = 'MFC '+str(i+1), linewidth = 2, linestyle = '--')
        ax2.tick_params(axis ='y', labelcolor = 'blue')
        ax.legend(fontsize = 'large')
        plt.title(label=title, fontsize = 'xx-large')
        plt.show()
    else:
        # Create lists for plotting
        plot_data = []
        for i in range(ws['A1'].value):  # Assuming 4 MFCs
            for j in range(len(current_density_average)):  # Assuming 13 data points
                plot_data.append((current_density_average[j], voltage[i][j], power_density[i][j], f'MFC {i+1}'))

        # Create the plot
        plot = (
            ggplot()
            + geom_line(aes(x=[p[0] for p in plot_data], y=[p[1] for p in plot_data], color=[p[3] for p in plot_data]))
            + geom_line(aes(x=[p[0] for p in plot_data], y=[p[2] for p in plot_data], linetype=[p[3] for p in plot_data]), size=1.2)
            + scale_color_manual(values=['red', 'green', 'blue', 'purple'])
            + scale_linetype_manual(values=['dashed', 'dotted', 'solid', 'dashdot'])
            + labs(x='Current density (mA/m^2)', y='Voltage (V) / Power Density (mW/m^2)', title='MFC Performance')
            + theme_minimal()
        )

        # print(plot)
        plot.draw()
    
    if save==True:
        if plot_matplotlib==True:
            fig.savefig('polarization_curve.png')
        else:
            plot.save('polarization_curve.png')

generate_polarization_curve(x, title = 'Polarization curve', plot_nine = True, plot_matplotlib = True, save=False)